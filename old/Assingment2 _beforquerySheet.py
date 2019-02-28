############################################################
#
# Assingment 2
#
#
#
#
############################################################

import numpy as np
import matplotlib.pyplot as plt
from pandas import DataFrame
from pandas import ExcelFile
import openpyxl
import math

def read_excel_sheet1(excelfile):
    from pandas import read_excel

    return (read_excel(excelfile)).values


def read_excel_range(excelfile, sheetname="Sheet1", startrow=1, endrow=1, startcol=1, endcol=1):
    from pandas import read_excel

    values = (read_excel(excelfile, sheetname, header=None)).values
    return values[startrow-1:endrow, startcol-1:endcol]


def read_excel(excelfile, **args):
    if args:
        data = read_excel_range(excelfile, **args)
    else:
        data = read_excel_sheet1(excelfile)

    if data.shape == (1, 1):
        return data[0, 0]
    elif (data.shape)[0] == 1:
        return data[0]
    else:
        return data


def get_sheet_names(excelfile):
    return (ExcelFile(excelfile)).sheet_names


def build_2D_histgram_classifier(gender_data, BINS, min_height, max_height, min_handspan, max_handspan):

    # get an height_binx*handspan_bins array of zero
    # Histogram must be square, use height_bins=22 instead of using handspan_bins
    # histogram_data = np.zeros((height_bins, handspan_bins), dtype='int32')
    histogram_data = np.zeros((BINS, BINS), dtype='int32')

    height_indices = (np.round(((BINS - 1) * (np.array(gender_data[:, 1], dtype=float) -
                                                     min_height) / (max_height - min_height)))).astype('int32')

    handspan_indices = (np.round(((BINS - 1) * (np.array(gender_data[:, 2], dtype=float) -
                                                         min_handspan) / (max_handspan - min_handspan)))).astype('int32')

#    print('height_indices->{}'.format(height_indices))
#    print('handspan_indices->{}'.format(handspan_indices))

    # count indices in each bin
    # for i, binindex in enumerate(height_indices):
    #   height_histo[binindex] += 1
    for i in range(len(height_indices)):
        histogram_data[height_indices[i], handspan_indices[i]] += 1

 #    print('histogram_data->{}'.format(histogram_data))

    return histogram_data

def debug_print(explanation, data ):
    print('{} -> {}'.format(explanation, data))


def write_histogram_excel(excel_file, sheet_name, histo_data, BINS, min_height, max_height, min_handspan, max_handspan):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # write max, min, bins data
    worksheet['B1'] = min_height
    worksheet['B2'] = max_height
    worksheet['B3'] = min_handspan
    worksheet['B4'] = max_handspan
    worksheet['B6'] = str(BINS) + 'x' + str(BINS)

    for i in range(BINS):
        for j in range(BINS):
            worksheet.cell(row=7+i, column=2+j, value=histo_data[i, j])

    wb.save(excel_file)


def get_probability_by_histogram(height, handspan, histo_female, histo_male, BINS, min_height, max_height, min_handspan, max_handspan):

    # get bins index in height and handspan
    height_bins_index = (np.round((BINS - 1) * (height-min_height) / (max_height - min_height))).astype('int32')
    handspan_bins_index = (np.round((BINS - 1) * (handspan-min_handspan) / (max_handspan - min_handspan))).astype('int32')

#    debug_print('height_bins_index',height_bins_index)
#    debug_print('handspan_bins_index', handspan_bins_index)

    # get count by gender
    count_female = histo_female[height_bins_index, handspan_bins_index]
    count_male = histo_male[height_bins_index, handspan_bins_index]

    # get Probability and gender
    if count_female + count_male == 0:
        # No Data
        probability = 0
        gender = 'NaN'

    elif count_female > count_male:
        # return femail probability
        probability = count_female / (count_female + count_male)
        gender = 'Female'

    else:
        probability = count_male / (count_female + count_male)
        gender = 'Male'

    return [gender, probability]


def draw_3d_histgram_2d_data(female_data, male_data, BINS, min_height, min_handspan):
    # Now this function is not used.

    # This import is not used directly, but it is used the column including projection='3d'
    from mpl_toolkits.mplot3d import Axes3D

    # plt.figure() function creates a canvas of charts( default 8*6 inches)
    # fig = plt.figure()
    # axis_1 = fig.add_subplot(111, projection='3d')

    np.random.seed(19680801)

    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    x, y = np.random.rand(2, 100) * 4
#    print('x->{}'.format(x))

#    print('female_data[1, ]->{}'.format(female_data[1, ]))

    hist, xedges, yedges = np.histogram2d(female_data[:, 1], female_data[:, 2], bins=21,
                                          range=[[min_height, min_height+21], [min_handspan, min_handspan+21]])

#    print('hist->{}'.format(hist))
#    print('xedges->{}'.format(xedges))
#    print('yedges->{}'.format(yedges))

    # Construct arrays for the anchor positions of the 16 bars.
    xpos, ypos = np.meshgrid(xedges[:-1], yedges[:-1], indexing="ij")
    xpos = xpos.ravel()
    ypos = ypos.ravel()
    zpos = 0

    # Construct arrays with the dimensions for the 16 bars.
    dx = dy = 0.5 * np.ones_like(zpos)
    dz = hist.ravel()

    ax.bar3d(xpos, ypos, zpos, dx, dy, dz, color='b', zsort='average')

    plt.show()


def get_count_by_bayesian(target_array, height_array, handspan_array):
    mean_height = height_array.mean()
    mean_handspan = handspan_array.mean()

    # variance
    #
    # height_variance = np.sum(np.exp(height_array-mean_height, 2))/len(height_array)
    # height_variance = math.pow(np.std(height_array), 2)
    # handspan_variance = math.pow(np.std(handspan_array), 2)

    #covariance = np.sum((height_array-mean_height)*(handspan_array-mean_handspan))/len(height_array)

    cov_array = np.vstack([height_array, handspan_array])
    #print('cov_array->{}'.format(cov_array))

    # get Covariance Matrix. bias=False means to get Unbiased Covariance
    cov_matrix = np.cov(cov_array.astype(float), bias=False)

    # print('cov_matrix->{}'.format(cov_matrix))

    # target_array = np.array([68, 22])

    diff_array = np.array([[target_array[0]-mean_height, target_array[1]-mean_handspan]])

    # get estimated count
    exp_num = np.exp(-(np.matmul(np.matmul(diff_array, np.linalg.inv(cov_matrix)), diff_array.T)/2))
    estimated_count = len(height_array) * (1 / ((2 * math.pi) * math.sqrt(np.linalg.det(cov_matrix))) * exp_num)

    # print('exp_num->{}'.format(exp_num))
    # print('np.linalg.det(cov_matrix)->{}'.format(np.linalg.det(cov_matrix)))

    # print('exp_num->{}'.format(exp_num))
    # print('estimated_count->{}'.format(estimated_count))

    return estimated_count


def write_querysheet_excel_bayes(excel_file, query, height_array_female, handspan_array_female, height_array_male, handspan_array_male):

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb['Queries']

    # query is 4 vectors in Query sheet in the Excel file
    for i, query_v in enumerate(query):
        print('query_v->{}'.format(query_v))
        female_count = get_count_by_bayesian(query_v, height_array_female, handspan_array_female).item()
        male_count = get_count_by_bayesian(query_v, height_array_male, handspan_array_male).item()

        if female_count > male_count:
            gender = 'Female'
            probability = female_count / (female_count + male_count)

        else:
            gender = 'Male'
            probability = male_count / (female_count + male_count)

        # write Gender and Probability on Excel Sheet
        worksheet['E'+str(3+i)] = gender
        worksheet['F'+str(3+i)] = probability

    wb.save(excel_file)

def write_querysheet_excel_histogram(excel_file, prpbability_array):
    # prpbability_array contains [[gender,probability]*row of query] matrix

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb['Queries']

    for i in range(len(query)):
        [gender, probability] = get_probability_by_histogram(query[i, 0], query[i, 1], histo_female, histo_male,
                                                            BINS, min_height, max_height,
                                                            min_handspan, max_handspan)

        worksheet['C' + str(3 + i)] = gender
        worksheet['D' + str(3 + i)] = probability

        # query is 4 vectors in Query sheet in the Excel file
    for i, probability_vector in enumerate(prpbability_array):
        #print(i)
        print(probability_vector[0,1])
        # write Gender and Probability on Excel Sheet
        #worksheet['C'+str(3+i)] = probability_vector[i, 0]
        #worksheet['D'+str(3+i)] = probability_vector[i,1]

    wb.save(excel_file)


def write_bayesian_excel(excel_file, height_array_female, handspan_array_female, height_array_male, handspan_array_male):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb['Bayesian']

    # write means
    worksheet['C1'] = height_array_female.mean()
    worksheet['D1'] = handspan_array_female.mean()

    worksheet['C2'] = height_array_male.mean()
    worksheet['D2'] = handspan_array_male.mean()

    # female Covariance Matrix
    cov_array_female = np.vstack([height_array_female, handspan_array_female])
    cov_matrix_female = np.cov(cov_array_female.astype(float), bias=False)

    worksheet['C4'] = cov_matrix_female[0, 0]
    worksheet['D4'] = cov_matrix_female[0, 1]
    worksheet['C5'] = cov_matrix_female[1, 0]
    worksheet['D5'] = cov_matrix_female[1, 1]

    # male Covariance Matrix
    cov_array_male = np.vstack([height_array_male, handspan_array_male])
    cov_matrix_male = np.cov(cov_array_male.astype(float), bias=False)

    worksheet['C6'] = cov_matrix_male[0, 0]
    worksheet['D6'] = cov_matrix_male[0, 1]
    worksheet['C7'] = cov_matrix_male[1, 0]
    worksheet['D7'] = cov_matrix_male[1, 1]

    # sample size
    worksheet['C9'] = len(height_array_female)
    worksheet['C10'] = len(height_array_male)

    wb.save(excel_file)


def write_histogram_by_bayesian_excel(excel_file, sheet_name, BINS, height_array, handspan_array, min_height, max_height, min_handspan, max_handspan):

    histogram_data = np.zeros((BINS, BINS), dtype='int32')

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # write max, min, bins data
    worksheet['B1'] = min_height
    worksheet['B2'] = max_height
    worksheet['B3'] = min_handspan
    worksheet['B4'] = max_handspan
    worksheet['B6'] = str(BINS) + 'x' + str(BINS)

    # get scale that one cell has.
    constant_height = (max_height-min_height)/(BINS)
    constant_handspan = (max_handspan-min_handspan)/(BINS)

    for i in range(BINS):
        for j in range(BINS):
            # calculate query of average height and handspan in the cell in the loop
            query_avg = np.array([min_height+(((i*constant_height)+(i*constant_height+constant_height))/2),
                                min_handspan+(((j*constant_handspan)+(j*constant_handspan+constant_handspan))/2)])

            # get count by Bayes classifier by using average height and handspan in the cell
            count_bayes_avg = get_count_by_bayesian(query_avg, height_array, handspan_array).item()

            # get count in the cell that has square, so multiply height and handspan
            count_bayes = count_bayes_avg*constant_height*constant_handspan

            # write the count to Excel. the count is rounded up in decimal point 3rd place.
            worksheet.cell(row=7+i, column=2+j, value=np.round(count_bayes, 3))

    wb.save(excel_file)


######################################################################################################
#
#   Main
#
#
######################################################################################################
excel_file=r"C:/Users/uskya/OneDrive/document_usk/UCSC/02_SecondQuater/Introduction_To_MachineLearning_DataMining/Assignment/2_Assignment/Assignment_2_Data_and_Template.xlsx"

sheets = get_sheet_names(excel_file)
# print("sheet -> {}".format(sheets))
data = read_excel(excel_file)

# print("full data of handspan -> {}".format(data[:,2]))

# MIN and MAX by using full data
min_height = np.amin(data[:, 1])
max_height = np.amax(data[:, 1])

min_handspan = np.amin(data[:, 2])
max_handspan = np.amax(data[:, 2])

# BINS by using full data
# BINS = math.log2(sampling size=167)+1 and round up to integer
BINS = 8

female_data = data[data[:, 0] == 'Female']
male_data = data[data[:, 0] == 'Male']

min_height_female = np.amin(female_data[:, 1])
max_height_female = np.amax(female_data[:, 1])

min_handspan_female = np.amin(female_data[:, 2])
max_handspan_female = np.amax(female_data[:, 2])

mean_height_female = (female_data[:, 1]).mean()
mean_handspan_female = (female_data[:, 2]).mean()

# print("min_height -> {}".format(min_height))
# print("max_height -> {}".format(max_height))

# print("min_handspan -> {}".format(min_handspan))
# print("max_handspan -> {}".format(max_handspan))

# print("min_height_female-> {}".format(min_height_female))
# print("max_height_female -> {}".format(max_height_female))

# print("min_handspan_female -> {}".format(min_handspan_female))
# print("max_handspan_female -> {}".format(max_handspan_female))

# get 22*22 matrix histogram data by gender
histo_female = build_2D_histgram_classifier(female_data, BINS, min_height, max_height, min_handspan, max_handspan)
histo_male = build_2D_histgram_classifier(male_data, BINS, min_height, max_height, min_handspan, max_handspan)

# write female histogram data to Excel. (The histogram must be square, only height_bins is used.)
write_histogram_excel(excel_file, 'Female Histogram', histo_female, BINS, min_height, max_height, min_handspan, max_handspan)
write_histogram_excel(excel_file, 'Male Histogram', histo_male, BINS, min_height, max_height, min_handspan, max_handspan)

# get query from 'Queries' sheet. query is the matrix of (height,handspan)*n rows
query = read_excel_range(excel_file,
                  sheetname='Queries',
                  startrow=3,
                  endrow=6,
                  startcol=1,
                  endcol=2).astype(float)

# debug_print('query', query)

# write probability of query by gender
probability_histo = np.array((["Female", 0.0],["Female", 0.0],["Female", 0.0],["Female", 0.0]))
for i in range(len(query)):
    probability_histo[i] = get_probability_by_histogram(query[i, 0], query[i, 1], histo_female, histo_male,
                                                         BINS, min_height, max_height,
                                                         min_handspan, max_handspan)
    #print(type(probability_histo))

    # debug_print('gender', gender)
    # debug_print('probability', probability)
# debug_print('probability_histo', probability_histo)

write_querysheet_excel_histogram(excel_file, query[i, 1], histo_female, histo_male,
                                                         BINS, min_height, max_height,
                                                         min_handspan, max_handspan)


# Draw 3D Histogram by using height and hand span data
# draw_3d_histgram_2d_data(female_data, male_data, height_bins, min_height, min_handspan)

# ####################################################################################################
# Bayesian
#
######################################################################################################
# wirte Bayesian sheet data to Excel
write_bayesian_excel(excel_file, female_data[:, 1], female_data[:, 2], male_data[:, 1], male_data[:, 2])

# wirte Queries sheet data to Excel
write_querysheet_excel_bayes(excel_file, query, female_data[:, 1], female_data[:, 2], male_data[:, 1], male_data[:, 2])

# write 'Reconstructed Female Histogram' sheet in Excel sheet
write_histogram_by_bayesian_excel(excel_file, 'Reconstructed Female Histogram', BINS, female_data[:, 1], female_data[:, 2],
                                  min_height, max_height, min_handspan, max_handspan)

# write 'Reconstructed Male Histogram' sheet in Excel sheet
write_histogram_by_bayesian_excel(excel_file, 'Reconstructed Male Histogram', BINS, male_data[:, 1], male_data[:, 2],
                                  min_height, max_height, min_handspan, max_handspan)

# TEST CODE
target_array = np.array([68, 22])
female_count = get_count_by_bayesian(target_array, female_data[:, 1], female_data[:, 2])
male_count = get_count_by_bayesian(target_array, male_data[:, 1], male_data[:, 2])

probability_female = female_count/(female_count+male_count)

print("female_count -> {}".format(female_count))
print("probability_female -> {}".format(probability_female))


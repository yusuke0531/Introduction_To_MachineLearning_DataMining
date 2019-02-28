import numpy as np
import matplotlib.pyplot as plt
from pandas import DataFrame


def readExcelSheet1(excelfile):
    from pandas import read_excel

    return (read_excel(excelfile)).values


def readExcelRange(excelfile, sheetname="Sheet1", startrow=1, endrow=1, startcol=1, endcol=1):
    from pandas import read_excel

    values = (read_excel(excelfile, sheetname, header=None)).values
    return values[startrow-1:endrow, startcol-1:endcol]


def readExcel(excelfile, **args):
    if args:
        data = readExcelRange(excelfile, **args)
    else:
        data = readExcelSheet1(excelfile)

    if data.shape == (1, 1):
        return data[0, 0]
    elif (data.shape)[0] == 1:
        return data[0]
    else:
        return data


def writeExcelData(x, excelfile, sheetname, startrow, startcol):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df=DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname, startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()


def writeExcelData2FullDataSheet(excelfile, data, female_bins, male_bins):
    from pandas import DataFrame, ExcelWriter
    import openpyxl
    import math

    female_data = data[data[:, 2] == 'Female']
    male_data = data[data[:, 2] == 'Male']
    print('Female data->{}'.format(female_data))
    print('Male data->{}'.format(male_data))

    female_array = np.array(female_data[:, 0] * 12 + female_data[:, 1], dtype=float)
    male_array = np.array(male_data[:, 0] * 12 + male_data[:, 1], dtype=float)
    all_array = np.array(data[:, 0] * 12 + data[:, 1], dtype=float)

    # Height inches MIN and MAX
    all_min = np.amin(all_array)
    all_max = np.amax(all_array)

    print('female_array->{}'.format(female_array))
    print('male_array->{}'.format(male_array))

    print('MIN->{}'.format(all_min))
    print('MAX->{}'.format(all_max))

    print('Female Mean->{}'.format(female_array.mean()))
    print('Male Mean->{}'.format(male_array.mean()))

    print('Female SD->{}'.format(np.std(female_array)))
    print('Male SD->{}'.format(np.std(male_array)))

    print('Female Size->{}'.format(len(female_array)))
    print('Male Size->{}'.format(len(male_array)))

    print('Feale bins Size->{}'.format(len(female_bins)))
    print('Male bins Size->{}'.format(len(male_bins)))


    # writer = ExcelWriter(excelfile, engine='xlsxwriter')
    # worksheet = writer.sheets['Classifiers - Full Data']

    # Write Excel Sheet
    wb = openpyxl.load_workbook(excelfile)
    worksheet = wb['Classifiers - Full Data']

    worksheet['B1']=all_min
    worksheet['B2']=all_max

    worksheet['C8']=female_array.mean()
    worksheet['C9']=male_array.mean()

    worksheet['C11'] = np.std(female_array)
    worksheet['C12'] = np.std(male_array)

    worksheet['C14'] = len(female_array)
    worksheet['C15'] = len(male_array)


    for i,b in enumerate(female_bins):
        worksheet.cell(row=5, column=3+i, value=female_bins[i])

    for i,b in enumerate(male_bins):
        worksheet.cell(row=6, column=3+i, value=male_bins[i])

    #  worksheet Queries
    worksheet_queries = wb['Queries']

    inch_index_55 = 55 - 52
    inch_index_60 = 60 - 52
    inch_index_65 = 65 - 52
    inch_index_70 = 70 - 52
    inch_index_75 = 75 - 52
    inch_index_80 = 80 - 52

    worksheet_queries['B3'] = 'Female'
    worksheet_queries['C3'] = female_bins[inch_index_55]/(female_bins[inch_index_55]+male_bins[inch_index_55])

    worksheet_queries['B4'] = 'Female'
    worksheet_queries['C4'] = female_bins[inch_index_60]/(female_bins[inch_index_60]+male_bins[inch_index_60])

    worksheet_queries['B5'] = 'Female'
    worksheet_queries['C5'] = female_bins[inch_index_65]/(female_bins[inch_index_65]+male_bins[inch_index_65])

    worksheet_queries['B6'] = 'Female'
    worksheet_queries['C6'] = female_bins[inch_index_70]/(female_bins[inch_index_70]+male_bins[inch_index_70])

    worksheet_queries['B7'] = 'Female'
    worksheet_queries['C7'] = female_bins[inch_index_75]/(female_bins[inch_index_75]+male_bins[inch_index_75])

    worksheet_queries['B8'] = 'Female'
    worksheet_queries['C8'] = female_bins[inch_index_80]/(female_bins[inch_index_80]+male_bins[inch_index_80])

    # Beyes
    # female_exp = np.exp(-1/2 * (55-female_array.mean()/np.std(female_array))**2)
    female_exp = math.exp(-(math.pow(55 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_55 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(55 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_55 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    print('Basyes Classifier ->{}'.format(female_55))
    print('Basyes Classifier ->{}'.format(male_55))
    print('Basyes Classifier F->{}'.format(np.round(female_55 / (female_55 + male_55), 10)))

    female_exp = math.exp(-(math.pow(60 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_60 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(60 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_60 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(65 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_65 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(65 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_65 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(70 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_70 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(70 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_70 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(75 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_75 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(75 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_75 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(80 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_80 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(80 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_80 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    worksheet_queries['D3'] = 'Female'
    worksheet_queries['E3'] = np.round(female_55 / (female_55 + male_55), 10)

    worksheet_queries['D4'] = 'Female'
    worksheet_queries['E4'] = np.round(female_60 / (female_60 + male_60), 10)

    worksheet_queries['D5'] = 'Female'
    worksheet_queries['E5'] = np.round(female_65 / (female_65 + male_65), 10)

    worksheet_queries['D6'] = 'Female'
    worksheet_queries['E6'] = np.round(female_70 / (female_70 + male_70), 10)

    worksheet_queries['D7'] = 'Female'
    worksheet_queries['E7'] = np.round(female_75 / (female_75 + male_75), 10)

    worksheet_queries['D8'] = 'Female'
    worksheet_queries['E8'] = np.round(female_80 / (female_80 + male_80), 10)

    wb.save(excelfile)

    return


def writeExcelData2PartialDataSheet(excelfile, data, female_bins, male_bins, all_min, all_max):
    from pandas import DataFrame, ExcelWriter
    import openpyxl
    import math

    female_data = data[data[:, 2] == 'Female']
    male_data = data[data[:, 2] == 'Male']
    print('Female data->{}'.format(female_data))
    print('Male data->{}'.format(male_data))

    female_array = np.array(female_data[:, 0] * 12 + female_data[:, 1], dtype=float)
    male_array = np.array(male_data[:, 0] * 12 + male_data[:, 1], dtype=float)

    print('female_array->{}'.format(female_array))
    print('male_array->{}'.format(male_array))

    print('MIN->{}'.format(all_min))
    print('MAX->{}'.format(all_max))

    print('Female Mean->{}'.format(female_array.mean()))
    print('Male Mean->{}'.format(male_array.mean()))

    print('Female SD->{}'.format(np.std(female_array)))
    print('Male SD->{}'.format(np.std(male_array)))

    print('Female Size->{}'.format(len(female_array)))
    print('Male Size->{}'.format(len(male_array)))

    print('Feale bins Size->{}'.format(len(female_bins)))
    print('Male bins Size->{}'.format(len(male_bins)))

    # Write Excel Sheet
    wb = openpyxl.load_workbook(excelfile)
    worksheet = wb['Classifiers - Partial Data']

    worksheet['B1']=all_min
    worksheet['B2']=all_max

    worksheet['C8']=female_array.mean()
    worksheet['C9']=male_array.mean()

    worksheet['C11'] = np.std(female_array)
    worksheet['C12'] = np.std(male_array)

    worksheet['C14'] = len(female_array)
    worksheet['C15'] = len(male_array)


    for i,b in enumerate(female_bins):
        worksheet.cell(row=5, column=3+i, value=female_bins[i])

    for i,b in enumerate(male_bins):
        worksheet.cell(row=6, column=3+i, value=male_bins[i])

    #  worksheet Queries
    worksheet_queries = wb['Queries']

    inch_index_55 = 55 - 52
    inch_index_60 = 60 - 52
    inch_index_65 = 65 - 52
    inch_index_70 = 70 - 52
    inch_index_75 = 75 - 52
    inch_index_80 = 80 - 52

    worksheet_queries['B12'] = 'Female'
    if female_bins[inch_index_55]+male_bins[inch_index_55] == 0:
        worksheet_queries['C12'] = 0
    else :
        worksheet_queries['C12'] = female_bins[inch_index_55]/(female_bins[inch_index_55]+male_bins[inch_index_55])

    worksheet_queries['B13'] = 'Female'
    if female_bins[inch_index_60]+male_bins[inch_index_60] == 0:
        worksheet_queries['C13'] = 0
    else :
        worksheet_queries['C13'] = female_bins[inch_index_60]/(female_bins[inch_index_60]+male_bins[inch_index_60])

    worksheet_queries['B14'] = 'Female'
    if female_bins[inch_index_65]+male_bins[inch_index_65] == 0:
        worksheet_queries['C14'] = 0
    else :
        worksheet_queries['C14'] = female_bins[inch_index_65]/(female_bins[inch_index_65]+male_bins[inch_index_65])

    worksheet_queries['B15'] = 'Female'
    if female_bins[inch_index_70]+male_bins[inch_index_70] == 0:
        worksheet_queries['C15'] = 0
    else :
        worksheet_queries['C15'] = female_bins[inch_index_70]/(female_bins[inch_index_70]+male_bins[inch_index_70])

    worksheet_queries['B16'] = 'Female'
    if female_bins[inch_index_75]+male_bins[inch_index_75] == 0:
        worksheet_queries['C16'] = 0
    else :
        worksheet_queries['C16'] = female_bins[inch_index_75]/(female_bins[inch_index_75]+male_bins[inch_index_75])

    worksheet_queries['B17'] = 'Female'
    if female_bins[inch_index_80]+male_bins[inch_index_80] == 0:
        worksheet_queries['C17'] = 0
    else :
        worksheet_queries['C17'] = female_bins[inch_index_80]/(female_bins[inch_index_80]+male_bins[inch_index_80])

    # Beyes
    # female_exp = np.exp(-1/2 * (55-female_array.mean()/np.std(female_array))**2)
    female_exp = math.exp(-(math.pow(55 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_55 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(55 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_55 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    print('Basyes Classifier ->{}'.format(female_55))
    print('Basyes Classifier ->{}'.format(male_55))
    print('Basyes Classifier F->{}'.format(np.round(female_55 / (female_55 + male_55), 10)))

    female_exp = math.exp(-(math.pow(60 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_60 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(60 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_60 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(65 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_65 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(65 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_65 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(70 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_70 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(70 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_70 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(75 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_75 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(75 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_75 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    female_exp = math.exp(-(math.pow(80 - female_array.mean(), 2) / (2 * math.pow(np.std(female_array), 2))))
    female_80 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(female_array)) * female_exp)

    male_exp = math.exp(-(math.pow(80 - male_array.mean(), 2) / (2 * math.pow(np.std(male_array), 2))))
    male_80 = len(male_array) * (1 / (math.sqrt(2 * math.pi) * np.std(male_array)) * male_exp)

    worksheet_queries['D12'] = 'Female'
    worksheet_queries['E12'] = np.round(female_55 / (female_55 + male_55), 10)

    worksheet_queries['D13'] = 'Female'
    worksheet_queries['E13'] = np.round(female_60 / (female_60 + male_60), 10)

    worksheet_queries['D14'] = 'Female'
    worksheet_queries['E14'] = np.round(female_65 / (female_65 + male_65), 10)

    worksheet_queries['D15'] = 'Female'
    worksheet_queries['E15'] = np.round(female_70 / (female_70 + male_70), 10)

    worksheet_queries['D16'] = 'Female'
    worksheet_queries['E16'] = np.round(female_75 / (female_75 + male_75), 10)

    worksheet_queries['D17'] = 'Female'
    worksheet_queries['E17'] = np.round(female_80 / (female_80 + male_80), 10)

    wb.save(excelfile)

    return


def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names


def Build1DHistogramClassifier(X, T, B, xmin, xmax):

    # create an array of 0, Height-Female, male
    HF = np.zeros(B).astype('int32')
    HM = np.zeros(B).astype('int32')

    # bin indices
    binindices = (np.round(((B-1)*(X-xmin)/(xmax-xmin)))).astype('int32')

    print('binindices:{}'.format(binindices))

    for i,b in enumerate(binindices):
        if T[i] == 'Female':
            HF[b] += 1
        else:
            HM[b] += 1
    return [HF, HM]


def Apply1DHistogramClassifier(queries, HF, HM, xmin, xmax):
    B = np.alen(HF)
    binindices = np.clip((np.round(((B-1)*(queries-xmin)/(xmax-xmin)))).astype('int32'), 0, B-1)

    countF = HF[binindices]
    countM = HM[binindices]

    resultlabel = np.full(np.alen(binindices), "Indeterminate", dtype=object)
    resultprob = np.full(np.alen(binindices), np.nan, dtype=object)

    indicesF = countF > countM
    indicesM = countM > countF

    print('indicesF:{}'.format(indicesF))

    resultlabel[indicesF] = "Female"
    resultlabel[indicesM] = "Male"

    probF = countF/(countF+countM)
    probM = countM/(countF+countM)

    resultprob[indicesF] = probF[indicesF]
    resultprob[indicesM] = probM[indicesM]

    return resultlabel, resultprob


########### MAIN ######################
excelfile=r"C:/Users/uskya/OneDrive/document_usk/UCSC/02_SecondQuater/Introduction_To_MachineLearning_DataMining/Assignment/1_Assignment/Assignment_1_Data_and_Template.xlsx"

sheets = getSheetNames(excelfile)
sheets
data = readExcel(excelfile)

data_partial = readExcel(excelfile,sheetname="Data", startrow=2, endrow=51, startcol=1, endcol=3)

print('data_partial = {}'.format(data_partial))

# Feet to inches
X = np.array(data[:, 0]*12+data[:, 1], dtype=float)
x_partial = np.array(data_partial[:, 0]*12+data_partial[:, 1], dtype=float)

# gender
T = np.array([str(g) for g in data[:, 2]])
t_partial = np.array([str(g) for g in data_partial[:, 2]])

queries = (readExcel(excelfile,
                  sheetname='Classifiers - Full Data',
                  startrow=17,
                  endrow=17,
                  startcol=2,
                  endcol=7)).astype(float)

print('queries = {}'.format(queries))
print(data[2])

B = 32

# Height inches MIN and MAX
xmin = np.amin(X)
xmax = np.amax(X)

print(xmin)
print(xmax)

# Count HF and HM by Classifier
[HF, HM] = Build1DHistogramClassifier(X, T, B, xmin, xmax)

[HF_partial, HM_partial] = Build1DHistogramClassifier(x_partial, t_partial, B, xmin, xmax)

writeExcelData2FullDataSheet(excelfile, data, HF, HM)
writeExcelData2PartialDataSheet(excelfile, data_partial, HF_partial, HM_partial, xmin, xmax)

plt.figure(figsize=(10, 5))
opacity = 0.5

# numbers over a specified interval
[bincenters, binwidth] = np.linspace(xmin, xmax, num=B, retstep=True)

# matplotlib.pyplot.bar(left, height, width=0.8, bottom=None, hold=None, data=None, **kwargs )
rects1 = plt.bar(bincenters-(binwidth/2), HF, binwidth,
                 alpha=opacity,
                 color='pink',
                 edgecolor='black',
                 label='Female')

rects2 = plt.bar(bincenters+(binwidth/2), HM, binwidth,
                 alpha=opacity,
                 color='b',
                 edgecolor='black',
                 label='Male')

plt.xlabel('Height')
plt.ylabel('Count')
# Get or set the x-limits of the current tick locations and labels.
plt.xticks(bincenters, bincenters.astype('int32'), fontsize=10)
# Places a legend on the axes.
plt.legend()
plt.show()

# [resultHlabel, resultHprob] = Apply1DHistogramClassifier(queries, HF, HM, xmin, xmax)
#  print(DataFrame([resultHlabel, resultHprob]).T)
# The result produced upon executing the above statement is not shown in this document



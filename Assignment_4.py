############################################################
#
# Assingment 4
#
#
############################################################
import struct
import numpy as np
from array import array
import matplotlib.pyplot as plt
from pandas import ExcelFile
import openpyxl
import numpy.linalg
import math
import pandas as pd


def read_excel_sheet1(excelfile):
    from pandas import read_excel

    return (read_excel(excelfile)).values


def read_excel_range(excelfile, sheetname="Sheet1", startrow=1, endrow=1, startcol=1, endcol=1):
    from pandas import read_excel

    values = (read_excel(excelfile, sheetname, header=None)).values
    return values[startrow-1:endrow, startcol-1:endcol]


def read_excel(excelfile, **args):
    if args:
        data = read_excel_range(excelfile, **args)
    else:
        data = read_excel_sheet1(excelfile)

    if data.shape == (1, 1):
        return data[0, 0]
    elif (data.shape)[0] == 1:
        return data[0]
    else:
        return data


def write_excel_rows(excel_file, sheet_name, start_row, start_column, vector_data):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row, column=start_column + i, value=vector_data[i])

    wb.save(excel_file)


def write_excel_columns(excel_file, sheet_name, start_row, start_column, vector_data):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row + i, column=start_column, value=vector_data[i])

    wb.save(excel_file)


def write_excel_rows_columns(excel_file, sheet_name, start_row, start_column, matrix_data):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i, row in enumerate(matrix_data):
        for j in range(len(row)):
            worksheet.cell(row=start_row + i, column=start_column + j, value=row[j])

    wb.save(excel_file)


def get_T_6class(T_base, data_size):
    # build 6 class classifier
    T_6class = np.full((data_size, 6), -1)

    for i, t in enumerate(T_base):
        T_6class[i, t] = 1

    return T_6class


######################################################################################################
#
#   Main
#
#
######################################################################################################
path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\4_Assignment'
# excel_file=r"C:/Users/uskya/OneDrive/document_usk/UCSC/02_SecondQuater/Introduction_To_MachineLearning_DataMining/Assignment/4_Assignment/Assignment_4_Data_and_Template.xlsx"
excel_file = path+"\\Assignment_4_Data_and_Template.xlsx"

# sheets = get_sheet_names(excel_file)

ORIGINAL_LOAD_FLAG = True

if ORIGINAL_LOAD_FLAG:
    data = read_excel(excelfile=excel_file, sheetname="Training Data", startrow=2, endrow=6601, startcol=1, endcol=17)
    test_data = read_excel(excelfile=excel_file, sheetname="To be classified", startrow=5, endrow=54, startcol=1, endcol=15)

else:
    data = np.load(path + "\\training.npy")
    test_data = np.load(path + "\\testing.npy")

# save data at once
# np.save(path+"\\training", data)
# np.save(path+"\\testing", test_data)

data_size = len(data)

x1 = np.full((data_size, 1), 1)
Xa = np.array(np.column_stack((x1, data[:, 0:15])), dtype=np.float)
T_Binary = np.array(data[:, 15:16], dtype=np.float)
PIV_Xa = np.linalg.pinv(Xa)
w_binary = np.matmul(PIV_Xa, T_Binary)

print("T_array->{}".format(T_Binary.shape))
print("Xa.shape->{}".format(Xa.shape))
print("PIV_Xa.shape->{}".format(PIV_Xa.shape))
print("w->{}".format(np.ravel(w_binary).shape))

write_excel_columns(excel_file, "Classifiers", 5, 1, np.ravel(w_binary))

T_6class_org = np.array(data[:, 16:17], dtype=np.int)
T_6class = get_T_6class(T_6class_org, data_size)

w_6class = np.matmul(PIV_Xa, T_6class)
print("w_6class->{}".format(w_6class.shape))

# write w (6-class Classifier) to Excel
write_excel_rows_columns(excel_file, "Classifiers", 5, 5, w_6class)


################################################
#
# Testing
#
################################################
testdata_size = len(test_data)

x1_test = np.full((testdata_size, 1), 1)
Xa_test = np.array(np.column_stack((x1_test, test_data[:, 0:15])), dtype=np.float)

# calculate T values
T_Binary_test_org = np.matmul(Xa_test, w_binary)
T_Binary_test = np.where(T_Binary_test_org < 0, -1, 1)

T_6class_test_org = np.matmul(Xa_test, w_6class)
T_6class_test = np.argmax(T_6class_test_org, axis=1)
print("T_6class_test->{}".format(T_6class_test.shape))
write_excel_columns(excel_file, "Classifiers", 5, 1, np.ravel(w_binary))

T_test_result = np.column_stack((T_Binary_test, T_6class_test))
write_excel_rows_columns(excel_file, "6class", 2, 1, T_6class_test_org)


################################################
#
# Accuracy by using training data
#
################################################
T_Binary_train_org = np.matmul(Xa, w_binary)
T_Binary_train = np.where(T_Binary_train_org < 0, -1, 1)

T_6class_train_org = np.matmul(Xa, w_6class)
T_6class_train = np.argmax(T_6class_train_org, axis=1)

T_train_result = np.column_stack((T_Binary_train, T_6class_train))
# write_excel_rows_columns(excel_file, "Training Data", 2, 18, T_train_result)
# np.add.at()

#f1 score  = 2*ppv*sensivity/(ppy+sensivity)
# Binary Classifier Accuracy
binary_1_1 = np.sum((T_Binary == 1) & (T_Binary_train == 1))
binary_1_n1 = np.sum((T_Binary == 1) & (T_Binary_train == -1))
binary_n1_1 = np.sum((T_Binary == -1) & (T_Binary_train == 1))
binary_n1_n1 = np.sum((T_Binary == -1) & (T_Binary_train == -1))

binary_accuracy = (binary_1_1 + binary_n1_n1) / (binary_1_1 + binary_1_n1 + binary_n1_1 + binary_n1_n1)
binary_sensitivity = binary_1_1 / (binary_1_1 + binary_1_n1)
binary_specificity = binary_n1_n1 / (binary_n1_1 + binary_n1_n1)
binary_PPV = binary_1_1 / (binary_1_1 + binary_n1_1)

# write Excel
binary_classified = np.empty((2, 2))
binary_classified[0, 0] = binary_1_1
binary_classified[0, 1] = binary_1_n1
binary_classified[1, 0] = binary_n1_1
binary_classified[1, 1] = binary_n1_n1
write_excel_rows_columns(excel_file, "Performance", 10, 3, binary_classified)

binary_stats = np.empty((1, 4))
binary_stats[0, 0] = binary_accuracy
binary_stats[0, 1] = binary_sensitivity
binary_stats[0, 2] = binary_specificity
binary_stats[0, 3] = binary_PPV
write_excel_columns(excel_file, "Performance", 8, 7, np.ravel(binary_stats))

print("binary_1_1->{}".format(binary_1_1))
print("binary_1_n1->{}".format(binary_1_n1))
print("binary_n1_1->{}".format(binary_n1_1))
print("binary_n1_n1->{}".format(binary_n1_n1))
print("binary_accuracy->{}".format(binary_accuracy))


# 6 Class Classifier Accuracy
ones_array = np.ones((len(T_6class), 1))
T_6class_array = np.column_stack((T_6class_org, T_6class_train, ones_array))
df_6class = pd.DataFrame(T_6class_array, columns=["TRUTH", "PREDICT", "ONES"])

# count_6class contains each cells' count. some indices doesn't exist. count_6class <= 6*6
count_6class = np.array(df_6class.groupby(['TRUTH', 'PREDICT'], as_index=False).count(), dtype='int')

classfied_6class = np.zeros((6, 6), dtype='int')


for i in range(len(count_6class)):
    classfied_6class[count_6class[i, 0], count_6class[i, 1]] = count_6class[i, 2]

print("classfied_6class->{}".format(classfied_6class))

ppv = np.zeros((6, 1))

for i in range(len(classfied_6class)):
    ppv[i] = classfied_6class[i, i]/(classfied_6class[:, i].sum())
    print("ppv{}:{}".format(i, ppv[i]))

max_index = np.argmax(ppv)
min_index = np.argmin(ppv)

ppv_matrix = np.empty((2, 2))
ppv_matrix[0, 0] = np.round(ppv[max_index]*100, decimals=2)
ppv_matrix[0, 1] = max_index
ppv_matrix[1, 0] = np.round(ppv[min_index]*100, decimals=2)
ppv_matrix[1, 1] = min_index

write_excel_rows_columns(excel_file, "Performance", 19, 3, classfied_6class)
write_excel_rows_columns(excel_file, "Performance", 20, 12, ppv_matrix)

print("np.argmax(ppv):", np.argmax(ppv))
print("np.argmin(ppv):", np.argmin(ppv))


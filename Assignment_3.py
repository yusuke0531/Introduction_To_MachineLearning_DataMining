############################################################
#
# Assingment 3
#
#
#
#
############################################################
import struct
import numpy as np
from array import array
import matplotlib.pyplot as plt
from pandas import ExcelFile
import openpyxl
import numpy.linalg
import math


def load_mnist(dataset="training", selectedDigits=range(10),
               path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\3_Assignment'):
    # Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        fname_digits = path + '\\' + 'train-images.idx3-ubyte'
        fname_labels = path + '\\' + 'train-labels.idx1-ubyte'
    elif dataset == "testing":
        fname_digits = path + '\\' + 't10k-images.idx3-ubyte'
        fname_labels = path + '\\' + 't10k-labels.idx1-ubyte'
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    # Import digits data, Open read binary mode
    digitsFileObject = open(fname_digits, 'rb')

    # FileObject.read(16) read 16 bytes as binary data
    # bytes object (in binary mode). size is an optional numeric argument
    # ">IIII" means > Big Endian, I is unsigned int (4bytes)
    #
    magic_nr, size, rows, cols = struct.unpack(">IIII", digitsFileObject.read(16))

    # read remaining all data. array.array("B",data) "B" means (unsigned char in C) int in Python 1 byte
    digitsData = array("B", digitsFileObject.read())
    digitsFileObject.close()
    print("digitsData SIZE_>{}".format(len(digitsData)))
    print("magic_nr, size, rows, cols->{},{},{},{}".format(magic_nr, size, rows, cols))

    # Import label data
    labelsFileObject = open(fname_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labelsFileObject.read(8))

    labelsData = array("B", labelsFileObject.read())
    labelsFileObject.close()

    # Find indices of selected digits
    indices = [k for k in range(size) if labelsData[k] in selectedDigits]
    N = len(indices)

    # Create empty arrays for X and T
    X = np.zeros((N, rows * cols), dtype=np.uint8)
    T = np.zeros((N, 1), dtype=np.uint8)


    # Fill X from digitsdata
    # Fill T from labelsdata
    for i in range(N):
        # get digitsdata as array from data array
        X[i] = digitsData[indices[i] * rows * cols:(indices[i] + 1) * rows * cols]

        # get label data
        T[i] = labelsData[indices[i]]

    return X, T


def get_class_data(data, labels, number):
    rows = 28
    cols = 28
    # Find indices of selected digits
    print(labels)
    indices = [k for k in range(len(labels)) if int(labels[k]) == number]
    N = len(indices)

    # Create empty arrays for X and T
    X = np.zeros((N, rows * cols), dtype=np.uint8)

    for i in range(N):
        # get digitsdata as array from data array
        X[i] = data[indices[i]]

    return X


def get_p_array(data, labels, number):
    cols = 2

    # Find indices of selected digits
    indices = [k for k in range(len(labels)) if int(labels[k] == number)]
    N = len(indices)

    # Create empty arrays for X and T
    p = np.zeros((N, cols), dtype=np.float)

    for i in range(N):
        p[i][0] = data[indices[i], 0]
        p[i][1] = data[indices[i], 1]

    return p


def get_test_vector(X_data, T_labels, number):

    indices = [k for k in range(len(T_labels)) if int(T_labels[k] == number)]

    # return the first vector
    return X_data[indices[0]]


def write_excel_row(excel_file, sheet_name, start_row, start_column, vector_data):
    # query contains [[height,handspan]*row of query] matrix

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row, column=start_column + i, value=vector_data[i])

    wb.save(excel_file)


def write_positive_histogram_excel(excel_file, sheet_name, BINS, histogram_data, sample_size, mean_vec, max_p1, min_p1, max_p2, min_p2, cov_matrix):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # put the numbers
    worksheet['B6'] = sample_size
    worksheet['B9'] = mean_vec[0]
    worksheet['C9'] = mean_vec[1]

    # covariance matrix
    worksheet['B12'] = cov_matrix[0, 0]
    worksheet['C12'] = cov_matrix[0, 1]
    worksheet['B13'] = cov_matrix[1, 0]
    worksheet['C13'] = cov_matrix[1, 1]

    # min and max
    worksheet['B17'] = min_p1
    worksheet['C17'] = max_p1

    worksheet['B18'] = min_p2
    worksheet['C18'] = max_p2

    for i in range(BINS):
        for j in range(BINS):
            worksheet.cell(row=20+i, column=2+j, value=histogram_data[i, j])

    wb.save(excel_file)


def write_negative_histogram_excel(excel_file, sheet_name, BINS, histogram_data, sample_size, mean_vec, max_p1, min_p1, max_p2, min_p2, cov_matrix):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # put the numbers
    worksheet['B7'] = sample_size
    worksheet['B10'] = mean_vec[0]
    worksheet['C10'] = mean_vec[1]

    # covariance matrix
    worksheet['B14'] = cov_matrix[0, 0]
    worksheet['C14'] = cov_matrix[0, 1]
    worksheet['B15'] = cov_matrix[1, 0]
    worksheet['C15'] = cov_matrix[1, 1]

    # min and max are already written in the function of positive
    # worksheet['B17'] = min_p1
    # worksheet['C17'] = max_p1

    # worksheet['B18'] = min_p2
    # worksheet['C18'] = max_p2

    for i in range(BINS):
        for j in range(BINS):
            worksheet.cell(row=53+i, column=2+j, value=histogram_data[i, j])

    wb.save(excel_file)


def build_2D_histgram_classifier(p_array, BINS, max_p1, min_p1, max_p2, min_p2):

    histogram_data = np.zeros((BINS, BINS), dtype='int32')

    p1_indices = (np.round(((BINS - 1) * (np.array(p_array[:, 0], dtype=float) - min_p1) / (max_p1 - min_p1)))).astype('int32')

    p2_indices = (np.round(((BINS - 1) * (np.array(p_array[:, 1], dtype=float) - min_p2) / (max_p2 - min_p2)))).astype('int32')

    # count indices in each bin
    for i in range(len(p_array)):
        histogram_data[p1_indices[i], p2_indices[i]] += 1

    return histogram_data


def get_probability_by_histogram(p1, p2, positive_histogram, negative_histogram, BINS, max_p1, min_p1, max_p2, min_p2, positive, negative):

    p1_index = (np.round((BINS - 1) * (p1-min_p1) / (max_p1 - min_p1))).astype('int32')
    p2_index = (np.round((BINS - 1) * (p2-min_p2) / (max_p2 - min_p2))).astype('int32')

    # get count by digit
    count_positive = positive_histogram[p1_index, p2_index]
    count_negative = negative_histogram[p1_index, p2_index]

    # get Probability and digit
    if count_positive + count_negative == 0:
        # No Data
        probability = 0
        digits = 'NaN'

    elif count_positive > count_negative:
        # return femail probability
        probability = count_positive / (count_positive + count_negative)
        digits = positive

    else:
        probability = count_negative / (count_positive + count_negative)
        digits = negative

    return [digits, probability]


def get_count_by_bayesian(target_p1, target_p2, mean_p1, mean_p2, cov_p, sample_size):

    diff_array = np.array([[target_p1-mean_p1, target_p2-mean_p2]])

    # get estimated count
    exp_num = np.exp(-(np.matmul(np.matmul(diff_array, np.linalg.inv(cov_p)), diff_array.T)/2))
    estimated_count = sample_size * (1 / ((2 * math.pi) * math.sqrt(np.linalg.det(cov_p))) * exp_num)

    # array to one number
    return estimated_count.item()


def get_probability_bayesian(count_positive, count_negative, positive, negative):

    if(count_positive > count_negative):
        digit = positive
        probability = count_positive / (count_positive + count_negative)
    else:
        digit = negative
        probability = count_negative / (count_positive + count_negative)

    return [digit, probability]


def get_accuracy_histogram(p_test, T_test, positive_histogram, negative_histogram, BINS, max_p1, min_p1, max_p2, min_p2, positive, negative):
    # get accuracy of hitogram by using full testing data set

    accuracy = 0
    size = len(p_test)

    for i in range(size):
        [digit, probability] = get_probability_by_histogram(p_test[i][0], p_test[i][1],
                                                                          positive_histogram, negative_histogram, BINS,
                                                                          max_p1, min_p1, max_p2, min_p2, positive,
                                                                          negative )

        # count up if digit == label. T_test[i] returns ndarray.T_test[i, 0] is better than T_test[i] in this if statement.
        if digit == T_test[i, 0]:
            accuracy += 1

    return accuracy/size


def get_accuracy_bayesian(p_test, T_test, mean_p_positive, mean_p_negative, C_positive, C_negative, size_positive,
                          size_negative, positive, negative):
    # get accuracy of hitogram by using full testing data set

    accuracy = 0
    size = len(p_test)

    for i in range(size):
        count_positive = get_count_by_bayesian(p_test[i][0], p_test[i][1],
                                                              mean_p_positive[0],
                                                              mean_p_positive[1],
                                                              C_positive, size_positive)

        count_negative = get_count_by_bayesian(p_test[i][0], p_test[i][1],
                                                              mean_p_negative[0],
                                                              mean_p_negative[1],
                                                              C_negative, size_negative)

        [digit, probability] = get_probability_bayesian(count_positive, count_negative, positive, negative)

        # count up if digit == label.T_test[i] returns ndarray.T_test[i, 0] is better than T_test[i] in this statement.
        if digit == T_test[i, 0]:
            accuracy += 1

    return accuracy/size


def vectortoimg(v, show=True):
    plt.imshow(v.reshape(28, 28), interpolation='None', cmap='gray')
    plt.axis('off')
    if show:
        plt.show()


def draw_scatter_plot(C_positive, C_negative, mean_p_positive, mean_p_negative, positive_size, negative_size, positive, negative):

    Pp = np.random.multivariate_normal(mean_p_positive, C_positive, positive_size)
    Tp = np.full((positive_size,), positive, dtype=np.int)

    Pn = np.random.multivariate_normal(mean_p_negative, C_negative, negative_size)
    Tn = np.full((negative_size,), negative, dtype=np.int)

    P = np.concatenate((Pn, Pp))
    T = np.concatenate((Tn, Tp))

    # For best effect, points should not be drawn in sequence but in random order
    np.random.seed(0)
    randomorder = np.random.permutation(np.arange(len(T)))
#    randomorder = np.arange(len(T))

    # Set colors
    cols = np.zeros((len(T), 4))  # Initialize matrix to hold colors
    cols[T == positive] = [0, 1, 0, 0.25]  # Positive points are green (with opacity 0.25)
    cols[T == negative] = [1, 0, 0, 0.25]  # Negative points are red (with opacity 0.25)


    # Draw scatter plot
    fig = plt.figure()
    ax = fig.add_subplot(111, facecolor='black')
    ax.scatter(P[randomorder, 1], P[randomorder, 0], s=5, linewidths=0, facecolors=cols[randomorder, :], marker="o");
    ax.set_aspect('equal')

    plt.gca().invert_yaxis()
    plt.show()

######################################################
#
# Main
#
######################################################
path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\3_Assignment'
excel_file=path+'\Assignment_3_ Submission_Template.xlsx'

dataset_name = "training"
negative = 5
positive = 6
BINS = 32

ORIGINAL_LOAD_FLAG = False

if ORIGINAL_LOAD_FLAG:
    X_data, T_data = load_mnist(dataset=dataset_name, selectedDigits=[negative, positive])
    X_test, T_test = load_mnist(dataset="testing", selectedDigits=[negative, positive])
else:
    X_data = np.load(path + "\\training_image.npy")
    T_data = np.load(path + "\\training_label.npy")
    X_test = np.load(path + "\\testing_image.npy")
    T_test = np.load(path + "\\testing_label.npy")

print("LOAD DATA>{}")

digitsDataX = np.array(X_data)
labelDataT = np.array(T_data)

mean_vector_u = np.mean(digitsDataX, axis=0)
Z_array = digitsDataX - mean_vector_u
# If rowvar is True (default), then each row represents a variable, with observations in the columns. Otherwise,
#  the relationship is transposed: each column represents a variable, while the rows contain observations.
C_covariance = np.cov(Z_array, rowvar=False)

# get the eigenvalues and eigenvectors from covariance matrix(symmetric matrix.)
# The column v[:, i] is the normalized eigenvector corresponding to the eigenvalue w[i].
[eigenvalues, V_eigenvectors] = numpy.linalg.eigh(C_covariance)


print("variance eigenvalues 2->{}".format((eigenvalues[-1]+eigenvalues[-2]) / np.sum(eigenvalues)))

eigenvalues_ud = np.flipud(eigenvalues)
for i in range(0,100,10):

    print("variance eigenvalues {}->{}".format(i,np.sum(eigenvalues_ud[0:i]) / np.sum(eigenvalues_ud)))

# write mean, the first, the second eigenvectorsof data to Excel
write_excel_row(excel_file, 'Results', 2, 2, mean_vector_u)
write_excel_row(excel_file, 'Results', 3, 2, V_eigenvectors[:, -1])
write_excel_row(excel_file, 'Results', 4, 2, V_eigenvectors[:, -2])

P_principal_component = np.dot(Z_array, V_eigenvectors.T)
R_ = np.dot(P_principal_component, V_eigenvectors)
V_eigenvectors_2class = np.vstack([V_eigenvectors[:,-1], V_eigenvectors[:,-2]])

# get p_array
p_array = np.matmul(Z_array, V_eigenvectors_2class.T)

###################### Positive Class #################
p_positive = get_p_array(p_array, T_data, positive)

# get mean in columns
mean_p_positive = np.mean(p_positive, axis=0)
C_positive = np.cov(p_positive, rowvar=False)

###################### Negative Class #################
p_negative = get_p_array(p_array, T_data, negative)

# get mean in columns
mean_p_negative = np.mean(p_negative, axis=0)
C_negative = np.cov(p_negative, rowvar=False)

############# MAX and MIN ###############################
# get Max and Min in p1 and p2 positive and negative
max_p1 = np.max(p_array[:, 0])
min_p1 = np.min(p_array[:, 0])
max_p2 = np.max(p_array[:, 1])
min_p2 = np.min(p_array[:, 1])

############## HISTOGRAM #################
positive_histogram = build_2D_histgram_classifier(p_positive, BINS, max_p1, min_p1, max_p2, min_p2)
negative_histogram = build_2D_histgram_classifier(p_negative, BINS, max_p1, min_p1, max_p2, min_p2)

# write data and histogram
write_positive_histogram_excel(excel_file, 'Results', BINS, positive_histogram, len(p_positive), mean_p_positive,
                               max_p1, min_p1, max_p2, min_p2, C_positive)

write_negative_histogram_excel(excel_file, 'Results', BINS, negative_histogram, len(p_negative), mean_p_negative,
                               max_p1, min_p1, max_p2, min_p2, C_negative)


######################################################
#
# Testing data set
#
######################################################
# POSITIVE DATA XZPR
X_test_positive = get_test_vector(X_test, T_test, positive)
Z_test_positive = X_test_positive - mean_vector_u
p_test_positive = np.matmul(Z_test_positive, V_eigenvectors_2class.T)
r_test_positive = np.matmul(p_test_positive, V_eigenvectors_2class)
x1_test_positive = r_test_positive + mean_vector_u


# write test data to Excel
# use write_excel_row(excel_file, sheet_name, start_row, start_column, vector_data)
write_excel_row(excel_file, "Results", 88, 2, X_test_positive)
write_excel_row(excel_file, "Results", 89, 2, Z_test_positive)
write_excel_row(excel_file, "Results", 90, 2, p_test_positive)
write_excel_row(excel_file, "Results", 91, 2, r_test_positive)
write_excel_row(excel_file, "Results", 92, 2, x1_test_positive)

# NEGATIVE DATA XZPR
X_test_negative = get_test_vector(X_test, T_test, negative)
Z_test_negative = X_test_negative - mean_vector_u
p_test_negative = np.matmul(Z_test_negative, V_eigenvectors_2class.T)
r_test_negative = np.matmul(p_test_negative, V_eigenvectors_2class)
x1_test_negative = r_test_negative + mean_vector_u

# write test data to Excel
write_excel_row(excel_file, "Results", 94, 2, X_test_negative)
write_excel_row(excel_file, "Results", 95, 2, Z_test_negative)
write_excel_row(excel_file, "Results", 96, 2, p_test_negative)
write_excel_row(excel_file, "Results", 97, 2, r_test_negative)
write_excel_row(excel_file, "Results", 98, 2, x1_test_negative)


# write probability of POSITIVE DATA to Excel
[digit_positive, probability_positive] = get_probability_by_histogram(p_test_positive[0], p_test_positive[1], positive_histogram,
                                                     negative_histogram, BINS, max_p1, min_p1, max_p2, min_p2, positive, negative)

[digit_negative, probability_negative] = get_probability_by_histogram(p_test_negative[0], p_test_negative[1], positive_histogram,
                                                     negative_histogram, BINS, max_p1, min_p1, max_p2, min_p2, positive, negative)

# write probability by histogram to Excel
write_excel_row(excel_file, "Results", 102, 2, [positive])
write_excel_row(excel_file, "Results", 103, 2, [digit_positive, probability_positive])
write_excel_row(excel_file, "Results", 106, 2, [negative])
write_excel_row(excel_file, "Results", 107, 2, [digit_negative, probability_negative])

# target p array is positive
count_positivetarget_positive = get_count_by_bayesian(p_test_positive[0], p_test_positive[1], mean_p_positive[0], mean_p_positive[1],
                                       C_positive, len(p_positive))

count_positivetarget_negative = get_count_by_bayesian(p_test_positive[0], p_test_positive[1], mean_p_negative[0], mean_p_negative[1],
                                                      C_negative, len(p_negative))

[digit_positive_bayes, probability_positive_bayes] = get_probability_bayesian(count_positivetarget_positive, count_positivetarget_negative,
                                                                              positive, negative)

write_excel_row(excel_file, "Results", 104, 2, [digit_positive_bayes, probability_positive_bayes])

# target p array is negative
count_negativetarget_positive = get_count_by_bayesian(p_test_negative[0], p_test_negative[1], mean_p_positive[0], mean_p_positive[1],
                                       C_positive, len(p_positive))

count_negativetarget_negative = get_count_by_bayesian(p_test_negative[0], p_test_negative[1], mean_p_negative[0], mean_p_negative[1],
                                                      C_negative, len(p_negative))

[digit_negative_bayes, probability_negative_bayes] = get_probability_bayesian(count_negativetarget_positive, count_negativetarget_negative,
                                                                              positive, negative)

write_excel_row(excel_file, "Results", 108, 2, [digit_negative_bayes, probability_negative_bayes])


######################################################
#
# Check accuracy by using full Testing data set
#
######################################################
Z_test = X_test - mean_vector_u
p_test = np.matmul(Z_test, V_eigenvectors_2class.T)

hitogram_accuracy = get_accuracy_histogram(p_test, T_test, positive_histogram, negative_histogram, BINS,
                                           max_p1, min_p1, max_p2, min_p2, positive, negative)

bayesian_accuracy = get_accuracy_bayesian(p_test, T_test, mean_p_positive, mean_p_negative, C_positive, C_negative, len(p_positive),
                                          len(p_negative), positive, negative)

write_excel_row(excel_file, "Results", 111, 2, [hitogram_accuracy])
write_excel_row(excel_file, "Results", 112, 2, [bayesian_accuracy])


######################################################
#
# Draw Data as digit
#
######################################################
 # COMMENT OUT
# vectortoimg(X_data[0])
# vectortoimg(X_data[2])
# vectortoimg(Z_array[0])
# vectortoimg(Z_array[2])

# vectortoimg(X_test_positive)
# vectortoimg(Z_test_positive)
# vectortoimg(r_test_positive)
# vectortoimg(x1_test_positive)

# vectortoimg(X_test_negative)
# vectortoimg(Z_test_negative)
# vectortoimg(r_test_negative)
# vectortoimg(x1_test_negative)

######################################################
#
# Draw a scatter plot
#
######################################################
draw_scatter_plot(C_positive, C_negative, mean_p_positive, mean_p_negative, len(p_positive), len(p_negative), positive, negative)


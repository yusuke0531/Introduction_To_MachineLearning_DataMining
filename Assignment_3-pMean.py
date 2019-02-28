############################################################
#
# Assingment 3
#
#
#
#
############################################################
import struct
import numpy as np
from array import array
import matplotlib.pyplot as plt
from pandas import ExcelFile
import openpyxl
import numpy.linalg

def load_mnist(dataset="training", selectedDigits=range(10),
               path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\3_Assignment'):
    # Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        fname_digits = path + '\\' + 'train-images.idx3-ubyte'
        fname_labels = path + '\\' + 'train-labels.idx1-ubyte'
    elif dataset == "testing":
        fname_digits = path + '\\' + 't10k-images.idx3-ubyte'
        fname_labels = path + '\\' + 't10k-labels.idx1-ubyte'
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    # Import digits data, Open read binary mode
    digitsFileObject = open(fname_digits, 'rb')

    # FileObject.read(16) read 16 bytes as binary data
    # bytes object (in binary mode). size is an optional numeric argument
    # ">IIII" means > Big Endian, I is unsigned int (4bytes)
    #
    magic_nr, size, rows, cols = struct.unpack(">IIII", digitsFileObject.read(16))

    # read remaining all data. array.array("B",data) "B" means (unsigned char in C) int in Python 1 byte
    digitsData = array("B", digitsFileObject.read())
    digitsFileObject.close()
    print("digitsData SIZE_>{}".format(len(digitsData)))
    print("magic_nr, size, rows, cols->{},{},{},{}".format(magic_nr, size, rows, cols))

    # Import label data
    labelsFileObject = open(fname_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labelsFileObject.read(8))
    print("magic_nr, size->{},{}".format(magic_nr, size))

    labelsData = array("B", labelsFileObject.read())
    labelsFileObject.close()

    # Find indices of selected digits
    indices = [k for k in range(size) if labelsData[k] in selectedDigits]
    N = len(indices)
    print("indices->{}".format(indices))

    # Create empty arrays for X and T
    X = np.zeros((N, rows * cols), dtype=np.uint8)
    T = np.zeros((N, 1), dtype=np.uint8)



    # Fill X from digitsdata
    # Fill T from labelsdata
    for i in range(N):
        # get digitsdata as array from data array
        X[i] = digitsData[indices[i] * rows * cols:(indices[i] + 1) * rows * cols]

        # get label data
        T[i] = labelsData[indices[i]]

    return X, T


def get_class_data(data, labels, number):
    rows = 28
    cols = 28
    # Find indices of selected digits
    print(labels)
    indices = [k for k in range(len(labels)) if int(labels[k]) == number]
    N = len(indices)

    # Create empty arrays for X and T
    X = np.zeros((N, rows * cols), dtype=np.uint8)

    for i in range(N):
        # get digitsdata as array from data array
        X[i] = data[indices[i]]

    return X



def write_excel_row(excel_file, sheet_name, start_row, start_column, vector_data):
    # query contains [[height,handspan]*row of query] matrix

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row, column=start_column + i, value=vector_data[i])

    wb.save(excel_file)


def write_positive_histogram_excel(excel_file, sheet_name, BINS, p_array_positive, mean_vec, max_p1, min_p1, max_p2, min_p2, cov_matrix):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # put the numbers
    worksheet['B6'] = len(p_array_positive)
    worksheet['B9'] = mean_vec[0]
    worksheet['C9'] = mean_vec[1]

    # covariance matrix
    worksheet['B12'] = cov_matrix[0, 0]
    worksheet['C12'] = cov_matrix[0, 1]
    worksheet['B13'] = cov_matrix[1, 0]
    worksheet['C13'] = cov_matrix[1, 1]

    # min and max
    worksheet['B17'] = min_p1
    worksheet['C17'] = max_p1

    worksheet['B18'] = min_p2
    worksheet['C18'] = max_p2

    # get Histogram data
    histogram_data = build_2D_histgram_classifier(p_array_positive, BINS, max_p1, min_p1, max_p2, min_p2)

    for i in range(BINS):
        for j in range(BINS):
            worksheet.cell(row=20+i, column=2+j, value=histogram_data[i, j])

    wb.save(excel_file)


def build_2D_histgram_classifier(p_array, BINS, max_p1, min_p1, max_p2, min_p2):

    # get an height_binx*handspan_bins array of zero
    # Histogram must be square, use height_bins=22 instead of using handspan_bins
    # histogram_data = np.zeros((height_bins, handspan_bins), dtype='int32')
    histogram_data = np.zeros((BINS, BINS), dtype='int32')

    height_indices = (np.round(((BINS - 1) * (np.array(p_array[:, 0], dtype=float) -
                                              min_p1) / (max_p1 - min_p1)))).astype('int32')

    handspan_indices = (np.round(((BINS - 1) * (np.array(p_array[:, 1], dtype=float) -
                                                min_p2) / (max_p2 - min_p2)))).astype('int32')

    # count indices in each bin
    # for i, binindex in enumerate(height_indices):
    #   height_histo[binindex] += 1
    for i in range(len(height_indices)):
        histogram_data[height_indices[i], handspan_indices[i]] += 1

    return histogram_data


#################### MAIN ##################################
path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\3_Assignment'
excel_file=path+'\Assignment_3_ Submission_Template.xlsx'

dataset_name = "training"
negative = 5
positive = 6

BINS = 32

X_data, T_data = load_mnist(dataset=dataset_name, selectedDigits=[negative, positive])

digitsDataX = np.array(X_data)
digitsDataX_tx = digitsDataX.T
print("digitsDataX_tx[160]->{}".format(digitsDataX_tx[160]))
labelDataT = np.array(T_data)
labelDataT_tx = labelDataT.T

mean_vector_u = np.mean(digitsDataX, axis=0)
print("mean digitsDataX[]160->{}".format((np.mean(digitsDataX,axis=0))[160]))
print("mean digitsDataX_tx[]160->{}".format(np.mean(digitsDataX_tx[160])))
print("mean mean_vector[160]->{},{}".format(mean_vector_u[160], len(mean_vector_u)))

print("mean_vector.shape->{}".format(mean_vector_u.shape))

# write mean of data to Excel
write_excel_row(excel_file, 'Results', 2, 2, mean_vector_u)

Z_array = digitsDataX - mean_vector_u

# print("z_array->{}".format(z_array[160]))
# print("z_array mean->{}".format(np.mean(Z_array, axis=0)))

# If rowvar is True (default), then each row represents a variable, with observations in the columns. Otherwise,
#  the relationship is transposed: each column represents a variable, while the rows contain observations.
C_covariance = np.cov(Z_array, rowvar=False)

# get the eigenvalues and eigenvectors from covariance matrix(symmetric matrix.)
# The column v[:, i] is the normalized eigenvector corresponding to the eigenvalue w[i].
[eigenvalues, V_eigenvectors] = numpy.linalg.eigh(C_covariance)

# print("eigenvalues ->{}".format(eigenvalues))
# print("V_eigenvectors ->{}".format(V_eigenvectors.shape))
# print("V_eigenvectors[-1] ->{}".format(V_eigenvectors[:, -1]))

# write first, second eigenvectors to Excel
write_excel_row(excel_file, 'Results', 3, 2, V_eigenvectors[:,-1])
write_excel_row(excel_file, 'Results', 4, 2, V_eigenvectors[:,-2])


P_principal_component = np.dot(Z_array, V_eigenvectors.T)
R_ = np.dot(P_principal_component, V_eigenvectors)

V_eigenvectors_2class =np.vstack([V_eigenvectors[:,-1], V_eigenvectors[:,-2]])
print("V_eigenvectors_2_class ->{}".format(V_eigenvectors_2class))


###################### Positive Class #################
X_positive = get_class_data(X_data, T_data, positive)
#print(X_positive[0])
print("X_positive ->{}".format(X_positive.shape))

mean_X_positive = np.mean(X_positive, axis=0)
# print("mean_vector_u_posi ->{}".format(mean_vector_u_posi))

# X_positive - mean_vector_u gives correct Z_positive. use mean of full(5,6)data.
Z_positive = X_positive - mean_X_positive
# Z_positive = X_positive - mean_vector_u

# get p1,p2 array
p_positive = np.matmul(Z_positive, V_eigenvectors_2class.T)

print("p_positive ->{}".format(p_positive[0:20]))

# get mean in columns
mean_p_positive = np.mean(p_positive,axis=0)
print("mean_p_positive ->{}".format(mean_p_positive))

C_positive = np.cov(p_positive, rowvar=False)
print("C_positive ->{}".format(C_positive))

###################### Positive Class #################
X_negative = get_class_data(X_data, T_data, negative)
print("X_negative ->{}".format(X_negative.shape))

mean_X_negative = np.mean(X_negative, axis=0)
# print("mean_vector_u_posi ->{}".format(mean_vector_u_posi))

Z_negative = X_negative - mean_X_negative

# get p1,p2 array
p_negative = np.matmul(Z_negative, V_eigenvectors_2class.T)

print("p_negative ->{}".format(p_negative[0:20]))

# get mean in columns
mean_p_negative = np.mean(p_negative, axis=0)
print("mean_p_negative ->{}".format(mean_p_negative))

C_negative = np.cov(p_negative, rowvar=False)
print("C_negative ->{}".format(C_negative))

############# MAX and MIN ###############################
# get Max and Min in p1 and p2 positive and negative
p_all = np.vstack([p_positive,p_negative])
max_p1_posi = np.max(p_all[:, 0])
min_p1_posi = np.min(p_all[:, 0])
max_p2_posi = np.max(p_all[:, 1])
min_p2_posi = np.min(p_all[:, 1])

print("max_p1 ->{},min_p1 ->{},max_p2 ->{},min_p2 ->{}".format(max_p1_posi, min_p1_posi, max_p2_posi, min_p2_posi))

############## Write Data to Excel #################
# write data and histogram
write_positive_histogram_excel(excel_file, 'Results', BINS, p_positive, mean_p_positive, max_p1_posi, min_p1_posi, max_p2_posi, min_p2_posi, C_positive)





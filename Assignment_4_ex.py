############################################################
#
# Assingment 4 extra
# Assignment_4_Extra_Credit.xlsx
#
############################################################
import struct
import numpy as np
from pandas import ExcelFile
import openpyxl
import pandas as pd


def read_excel_sheet1(excelfile):
    from pandas import read_excel

    return (read_excel(excelfile)).values


def read_excel_range(excelfile, sheetname="Sheet1", startrow=1, endrow=1, startcol=1, endcol=1):
    from pandas import read_excel

    values = (read_excel(excelfile, sheetname, header=None)).values
    return values[startrow-1:endrow, startcol-1:endcol]


def read_excel(excelfile, **args):
    if args:
        data = read_excel_range(excelfile, **args)
    else:
        data = read_excel_sheet1(excelfile)

    if data.shape == (1, 1):
        return data[0, 0]
    elif (data.shape)[0] == 1:
        return data[0]
    else:
        return data


def write_excel_rows(excel_file, sheet_name, start_row, start_column, vector_data):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row, column=start_column + i, value=vector_data[i])

    wb.save(excel_file)


def write_excel_columns(excel_file, sheet_name, start_row, start_column, vector_data):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row + i, column=start_column, value=vector_data[i])

    wb.save(excel_file)


def write_excel_rows_columns(excel_file, sheet_name, start_row, start_column, matrix_data):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i, row in enumerate(matrix_data):
        for j in range(len(row)):
            worksheet.cell(row=start_row + i, column=start_column + j, value=row[j])

    wb.save(excel_file)


def get_T_6class(T_base, data_size):
    # build 6 class classifier
    T_6class = np.full((data_size, 6), -1)

    for i, t in enumerate(T_base):
        T_6class[i, t] = 1

    return T_6class


######################################################################################################
#
#   Main
#
#
######################################################################################################
path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\4_Assignment'
# excel_file=r"C:/Users/uskya/OneDrive/document_usk/UCSC/02_SecondQuater/Introduction_To_MachineLearning_DataMining/Assignment/4_Assignment/Assignment_4_Data_and_Template.xlsx"
excel_file = path+"\\Assignment_4_Data_and_Template.xlsx"
excel_file_extra = path+"\\Assignment_4_Extra_Credit.xlsx"
# sheets = get_sheet_names(excel_file)

ORIGINAL_LOAD_FLAG = False

if ORIGINAL_LOAD_FLAG:
    data = read_excel(excelfile=excel_file, sheetname="Training Data", startrow=2, endrow=6601, startcol=1, endcol=17)

else:
    data = np.load(path + "\\training.npy")

data_size = len(data)

x1 = np.full((data_size, 1), 1)
Xa = np.array(np.column_stack((x1, data[:, 0:15])), dtype=np.float)
T_Binary = np.array(data[:, 15:16], dtype=np.float)
PIV_Xa = np.linalg.pinv(Xa)
w_binary = np.matmul(PIV_Xa, T_Binary)

################################################
#
# Accuracy by using training data
#
################################################
Xa_4 = Xa[0:4]
T_Binary_bp = np.matmul(Xa_4, w_binary)
print("T_Binary_bp->{}".format(T_Binary_bp))

Xa_4_b1 = Xa_4
Xa_4_b1[0:4] = Xa_4_b1[0:4]
T_Binary_b1 = np.matmul(Xa_4_b1*0.1, w_binary)
print("T_Binary_b1->{}".format(T_Binary_b1))

Xa_4_b2 = Xa_4
T_Binary_b2 = np.matmul(Xa_4_b2*0.2, w_binary)
print("T_Binary_b2->{}".format(T_Binary_b2))

Xa_4_b3 = Xa_4
T_Binary_b3 = np.matmul(Xa_4_b3*0.3, w_binary)
print("T_Binary_b3->{}".format(T_Binary_b3))

Xa_4_b4 = Xa_4
T_Binary_b4 = np.matmul(Xa_4_b4/0.9, w_binary)
print("T_Binary_b4->{}".format(T_Binary_b4))

Xa_4_b5 = Xa_4
T_Binary_b5 = np.matmul(Xa_4_b5/0.8, w_binary)
print("T_Binary_b5->{}".format(T_Binary_b5))

T_Binary_all = np.column_stack((T_Binary_bp, T_Binary_b1, T_Binary_b2, T_Binary_b3,T_Binary_b4,T_Binary_b5))
print("T_Binary_all->{}".format(T_Binary_all))

write_excel_rows_columns(excel_file_extra, "Results", 11, 1, T_Binary_all)



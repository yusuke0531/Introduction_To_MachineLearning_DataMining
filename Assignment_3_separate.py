############################################################
#
# Assingment 3
#
#
#
#
############################################################
import struct
import numpy as np
from array import array
import matplotlib.pyplot as plt
from pandas import ExcelFile
import openpyxl
import numpy.linalg

def load_mnist(dataset="training", selectedDigits=range(10),
               path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\3_Assignment'):
    # Check training/testing specification. Must be "training" (default) or "testing"
    if dataset == "training":
        fname_digits = path + '\\' + 'train-images.idx3-ubyte'
        fname_labels = path + '\\' + 'train-labels.idx1-ubyte'
    elif dataset == "testing":
        fname_digits = path + '\\' + 't10k-images.idx3-ubyte'
        fname_labels = path + '\\' + 't10k-labels.idx1-ubyte'
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    # Import digits data, Open read binary mode
    digitsFileObject = open(fname_digits, 'rb')

    # FileObject.read(16) read 16 bytes as binary data
    # bytes object (in binary mode). size is an optional numeric argument
    # ">IIII" means > Big Endian, I is unsigned int (4bytes)
    #
    magic_nr, size, rows, cols = struct.unpack(">IIII", digitsFileObject.read(16))

    # read remaining all data. array.array("B",data) "B" means (unsigned char in C) int in Python 1 byte
    digitsData = array("B", digitsFileObject.read())
    digitsFileObject.close()
    print("digitsData SIZE_>{}".format(len(digitsData)))
    print("magic_nr, size, rows, cols->{},{},{},{}".format(magic_nr, size, rows, cols))

    # Import label data
    labelsFileObject = open(fname_labels, 'rb')
    magic_nr, size = struct.unpack(">II", labelsFileObject.read(8))
    print("magic_nr, size->{},{}".format(magic_nr, size))

    labelsData = array("B", labelsFileObject.read())
    labelsFileObject.close()

    # Find indices of selected digits
    indices = [k for k in range(size) if labelsData[k] in selectedDigits]
    N = len(indices)
    print("indices->{}".format(indices))

    # Create empty arrays for X and T
    X = np.zeros((N, rows * cols), dtype=np.uint8)
    T = np.zeros((N, 1), dtype=np.uint8)



    # Fill X from digitsdata
    # Fill T from labelsdata
    for i in range(N):
        # get digitsdata as array from data array
        X[i] = digitsData[indices[i] * rows * cols:(indices[i] + 1) * rows * cols]

        # get label data
        T[i] = labelsData[indices[i]]

    return X, T


def get_class_data(data, labels, number):
    rows = 28
    cols = 28
    # Find indices of selected digits
    print(labels)
    indices = [k for k in range(len(labels)) if int(labels[k]) == number]
    N = len(indices)

    # Create empty arrays for X and T
    X = np.zeros((N, rows * cols), dtype=np.uint8)

    for i in range(N):
        # get digitsdata as array from data array
        X[i] = data[indices[i]]

    return X


def get_p_array(data, labels, number):
    cols = 2

    # Find indices of selected digits
    print("data[20]->{}".format(data[0:20, :]))
    print("labels->{}".format(labels))

    indices = [k for k in range(len(labels)) if int(labels[k] == number)]
    N = len(indices)

    # Create empty arrays for X and T
    p = np.zeros((N, cols), dtype=np.float)

    for i in range(N):
        p[i][0] = data[indices[i], 0]
        p[i][1] = data[indices[i], 1]

    return p

def get_test_vector(X_data, T_labels, number):

    indices = [k for k in range(len(T_labels)) if int(T_labels[k] == number)]

    # return the first vector
    return X_data[indices[0]]

def write_excel_row(excel_file, sheet_name, start_row, start_column, vector_data):
    # query contains [[height,handspan]*row of query] matrix

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    for i in range(len(vector_data)):
        worksheet.cell(row=start_row, column=start_column + i, value=vector_data[i])

    wb.save(excel_file)


def write_positive_histogram_excel(excel_file, sheet_name, BINS, histogram_data, sample_size, mean_vec, max_p1, min_p1, max_p2, min_p2, cov_matrix):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # put the numbers
    worksheet['B6'] = sample_size
    worksheet['B9'] = mean_vec[0]
    worksheet['C9'] = mean_vec[1]

    # covariance matrix
    worksheet['B12'] = cov_matrix[0, 0]
    worksheet['C12'] = cov_matrix[0, 1]
    worksheet['B13'] = cov_matrix[1, 0]
    worksheet['C13'] = cov_matrix[1, 1]

    # min and max
    worksheet['B17'] = min_p1
    worksheet['C17'] = max_p1

    worksheet['B18'] = min_p2
    worksheet['C18'] = max_p2

    for i in range(BINS):
        for j in range(BINS):
            worksheet.cell(row=20+i, column=2+j, value=histogram_data[i, j])

    wb.save(excel_file)


def write_negative_histogram_excel(excel_file, sheet_name, BINS, histogram_data, sample_size, mean_vec, max_p1, min_p1, max_p2, min_p2, cov_matrix):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb[sheet_name]

    # put the numbers
    worksheet['B7'] = sample_size
    worksheet['B10'] = mean_vec[0]
    worksheet['C10'] = mean_vec[1]

    # covariance matrix
    worksheet['B14'] = cov_matrix[0, 0]
    worksheet['C14'] = cov_matrix[0, 1]
    worksheet['B15'] = cov_matrix[1, 0]
    worksheet['C15'] = cov_matrix[1, 1]

    # min and max are already written in the function of positive
    # worksheet['B17'] = min_p1
    # worksheet['C17'] = max_p1

    # worksheet['B18'] = min_p2
    # worksheet['C18'] = max_p2

    for i in range(BINS):
        for j in range(BINS):
            worksheet.cell(row=53+i, column=2+j, value=histogram_data[i, j])

    wb.save(excel_file)

def build_2D_histgram_classifier(p_array, BINS, max_p1, min_p1, max_p2, min_p2):

    histogram_data = np.zeros((BINS, BINS), dtype='int32')

    p1_indices = (np.round(((BINS - 1) * (np.array(p_array[:, 0], dtype=float) - min_p1) / (max_p1 - min_p1)))).astype('int32')

    p2_indices = (np.round(((BINS - 1) * (np.array(p_array[:, 1], dtype=float) - min_p2) / (max_p2 - min_p2)))).astype('int32')

    # count indices in each bin
    for i in range(len(p_array)):
        histogram_data[p1_indices[i], p2_indices[i]] += 1

    return histogram_data


def write_querysheet_excel_histogram(excel_file, query, histo_female, histo_male, BINS, min_height, max_height,
                                                            min_handspan, max_handspan):

    # query contains [[height,handspan]*row of query] matrix

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb['Queries']

    for i in range(len(query)):
        [gender, probability] = get_probability_by_histogram(query[i, 0], query[i, 1], histo_female, histo_male,
                                                            BINS, min_height, max_height,
                                                            min_handspan, max_handspan)

        worksheet['C' + str(3 + i)] = gender
        worksheet['D' + str(3 + i)] = probability

    wb.save(excel_file)


def get_probability_by_histogram(height, handspan, histo_female, histo_male, BINS, min_height, max_height, min_handspan, max_handspan):

    # get bins index in height and handspan
    height_bins_index = (np.round((BINS - 1) * (height-min_height) / (max_height - min_height))).astype('int32')
    handspan_bins_index = (np.round((BINS - 1) * (handspan-min_handspan) / (max_handspan - min_handspan))).astype('int32')

#    debug_print('height_bins_index',height_bins_index)
#    debug_print('handspan_bins_index', handspan_bins_index)

    # get count by gender
    count_female = histo_female[height_bins_index, handspan_bins_index]
    count_male = histo_male[height_bins_index, handspan_bins_index]

    # get Probability and gender
    if count_female + count_male == 0:
        # No Data
        probability = 0
        gender = 'NaN'

    elif count_female > count_male:
        # return femail probability
        probability = count_female / (count_female + count_male)
        gender = 'Female'

    else:
        probability = count_male / (count_female + count_male)
        gender = 'Male'

    return [gender, probability]


######################################################
#
# Main
#
######################################################
path=r'C:\Users\uskya\OneDrive\document_usk\UCSC\02_SecondQuater\Introduction_To_MachineLearning_DataMining\Assignment\3_Assignment'
excel_file=path+'\Assignment_3_ Submission_Template.xlsx'

dataset_name = "training"
negative = 5
positive = 6
BINS = 32

X_data, T_data = load_mnist(dataset=dataset_name, selectedDigits=[negative, positive])

digitsDataX = np.array(X_data)
labelDataT = np.array(T_data)

mean_vector_u = np.mean(digitsDataX, axis=0)
Z_array = digitsDataX - mean_vector_u
# If rowvar is True (default), then each row represents a variable, with observations in the columns. Otherwise,
#  the relationship is transposed: each column represents a variable, while the rows contain observations.
C_covariance = np.cov(Z_array, rowvar=False)

# get the eigenvalues and eigenvectors from covariance matrix(symmetric matrix.)
# The column v[:, i] is the normalized eigenvector corresponding to the eigenvalue w[i].
[eigenvalues, V_eigenvectors] = numpy.linalg.eigh(C_covariance)

# write mean, the first, the second eigenvectorsof data to Excel
write_excel_row(excel_file, 'Results', 2, 2, mean_vector_u)
write_excel_row(excel_file, 'Results', 3, 2, V_eigenvectors[:, -1])
write_excel_row(excel_file, 'Results', 4, 2, V_eigenvectors[:, -2])

P_principal_component = np.dot(Z_array, V_eigenvectors.T)
R_ = np.dot(P_principal_component, V_eigenvectors)
V_eigenvectors_2class = np.vstack([V_eigenvectors[:,-1], V_eigenvectors[:,-2]])

# get p_array
p_array = np.matmul(Z_array, V_eigenvectors_2class.T)

###################### Positive Class #################
X_positive = get_class_data(X_data, T_data, positive)
Z_positive = X_positive - mean_vector_u

p_positive = np.matmul(Z_positive, V_eigenvectors_2class.T)

# get mean in columns
mean_p_positive = np.mean(p_positive, axis=0)
C_positive = np.cov(p_positive, rowvar=False)

###################### Negative Class #################
p_negative = get_p_array(p_array, T_data, negative)

# get mean in columns
mean_p_negative = np.mean(p_negative, axis=0)
C_negative = np.cov(p_negative, rowvar=False)

############# MAX and MIN ###############################
# get Max and Min in p1 and p2 positive and negative
p_all = np.vstack([p_positive, p_negative])
max_p1 = np.max(p_all[:, 0])
min_p1 = np.min(p_all[:, 0])
max_p2 = np.max(p_all[:, 1])
min_p2 = np.min(p_all[:, 1])

############## HISTOGRAM #################
positive_histogram = build_2D_histgram_classifier(p_positive, BINS, max_p1, min_p1, max_p2, min_p2)
negative_histogram = build_2D_histgram_classifier(p_negative, BINS, max_p1, min_p1, max_p2, min_p2)

# write data and histogram
write_positive_histogram_excel(excel_file, 'Results', BINS, positive_histogram, len(p_positive), mean_p_positive,
                               max_p1, min_p1, max_p2, min_p2, C_positive)

write_negative_histogram_excel(excel_file, 'Results', BINS, negative_histogram, len(p_negative), mean_p_negative,
                               max_p1, min_p1, max_p2, min_p2, C_negative)


######################################################
#
# Testing data set
#
######################################################

X_test, T_test = load_mnist(dataset="testing", selectedDigits=[negative, positive])

# POSITIVE DATA XZPR
X_test_positive = get_test_vector(X_test, T_test, positive)
Z_test_positive = X_test_positive - mean_vector_u
p_test_positive = np.matmul(Z_test_positive, V_eigenvectors_2class.T)
r_test_positive = np.matmul(p_test_positive, V_eigenvectors_2class)
x1_test_positive = r_test_positive + mean_vector_u


# write test data to Excel
# use write_excel_row(excel_file, sheet_name, start_row, start_column, vector_data)
write_excel_row(excel_file, "Results", 88, 2, X_test_positive)
write_excel_row(excel_file, "Results", 89, 2, Z_test_positive)
write_excel_row(excel_file, "Results", 90, 2, p_test_positive)
write_excel_row(excel_file, "Results", 91, 2, r_test_positive)
write_excel_row(excel_file, "Results", 92, 2, x1_test_positive)

# NEGATIVE DATA XZPR
X_test_negative = get_test_vector(X_test, T_test, negative)
Z_test_negative = X_test_negative - mean_vector_u
p_test_negative = np.matmul(Z_test_negative, V_eigenvectors_2class.T)
r_test_negative = np.matmul(p_test_negative, V_eigenvectors_2class)
x1_test_negative = r_test_negative + mean_vector_u

# write test data to Excel
write_excel_row(excel_file, "Results", 94, 2, X_test_negative)
write_excel_row(excel_file, "Results", 95, 2, Z_test_negative)
write_excel_row(excel_file, "Results", 96, 2, p_test_negative)
write_excel_row(excel_file, "Results", 97, 2, r_test_negative)
write_excel_row(excel_file, "Results", 98, 2, x1_test_negative)


# write probability of POSITIVE DATA to Excel


